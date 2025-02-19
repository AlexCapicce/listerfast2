import cv2
import face_recognition
from flask import Flask, Response, render_template, request, redirect, url_for, flash, session, make_response, send_file
#from database import get_user_credentials
from database import *
from notificar_ausencias import *

#import pickle
import threading
from recognition import FaceRecognition
#from notificar_ausencias import exportar_registro_pdf
from openpyxl import Workbook
import os
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash
from werkzeug.security import generate_password_hash
import json
# Librerias para RF 
from flask import jsonify, Response
from functools import wraps
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen.canvas import Canvas  # Importación correcta
from reportlab.pdfgen import canvas

import numpy as np
import pandas as pd
from fpdf import FPDF
import io
#url_api = "http://127.0.0.1:5000/api/estudiantes"
app = Flask(__name__)
app.secret_key = "clave_secreta"

'''try:
    response = requests.get(url_api)
    response.raise_for_status()
    print("Conexión exitosa a la API. Datos obtenidos:")
    print(response.json())
except requests.exceptions.RequestException as e:
    print(f"Error al conectar con la API: {e}")'''

#iniciar_sheduler()

# Ejecutar el scheduler en un hilo separado
#thread = threading.Thread(target=iniciar_scheduler, daemon=True)
#thread.start()

@app.route("/")
def index():
    return render_template("login.html")
    #return render_template("index.html", cursos=cursos, materias=materias)

'''@app.route("/login", methods=["POST"])
def login():

    username = request.form["username"]
    password = request.form["password"]
    
    # Verificar credenciales del usuario
    user = get_user_credentials(username)
    
    if user and user["password"] == password:
        session["user"] = user["username"]
        flash("Inicio de sesión exitoso", "success")
        return redirect(url_for("dashboard"))
    else:
        flash("Usuario o contraseña incorrectos", "danger")
        return redirect(url_for("index"))'''

'''@app.route("/loginx", methods=['GET', 'POST'])
def loginx():
    username = request.form["username"]
    password = request.form["password"]
    
    # Verificar credenciales del usuario
    user = get_user_credentials(username)
    
    if user and user["password"] == password:
        session["user"] = user["username"]
        flash("Inicio de sesión exitoso", "success")
        return redirect(url_for("dashboard"))
    else:
        flash("Usuario o contraseña incorrectos", "danger")
        return redirect(url_for("index"))'''
    
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        correo = request.form['correo']
        contrasena = request.form['contrasena']
        hash_contrasena = generate_password_hash(contrasena)
        
        conexion=conectar_bd()
        cursor = conexion.cursor(dictionary=True)

        #hash_contrasena = generate_password_hash(contrasena1)

       # Consultar el usuario en la base de datos
        cursor.execute("SELECT * FROM usuario WHERE correo = %s", (correo,))
        usuario = cursor.fetchone()
        conexion.close()

        # Verificar si el usuario existe y si la contraseña es correcta
        if usuario and check_password_hash(usuario['contrasena'], contrasena):
            session['usuario_id'] = usuario['id_usuario']
            session['rol'] = usuario['id_rol']
            flash('Inicio de sesión exitoso', 'success')

            # Redirigir según el rol
            if usuario['id_rol'] == 1:  # Admin
                return redirect(url_for('dashboard_admin'))
            elif usuario['id_rol'] == 2:  # Docente
                return redirect(url_for('dashboard_docente'))
            elif usuario['id_rol'] == 3:  # Padre
                return redirect(url_for('dashboard_padre'))
            else:
                flash('Rol no permitido', 'danger')
                return redirect(url_for('login'))
            
        else:
            flash('Correo o contraseña incorrectos', 'danger')

    return render_template('login.html')

@app.route("/dashboard")
def dashboard():
    if "user" in session:
        return render_template("dashboard.html", user=session["user"])
    else:
        flash("Por favor, inicie sesión primero", "warning")
        return redirect(url_for("index"))

@app.route("/logout")
def logout():
    session.pop("user", None)
    flash("Sesión cerrada exitosamente", "info")
    return redirect(url_for("index"))

@app.route('/registro_asistencia', methods=['GET'])
def registro_asistencia():
    cursos = obtener_cursos()  # Función para obtener cursos desde la base de datos
    materias = obtener_materias1()  # Función para obtener materias desde la base de datos
    
    return render_template('registro_asistencia.html', cursos=cursos, materias=materias)

'''@app.route("/exportar-pdf")
def exportar_pdf():
    try:

        nombre_archivo = exportar_registro_pdf()
        if nombre_archivo:
            return send_file(nombre_archivo, as_attachment=True)
    
        
    except Exception as e:
        print(f"Error al generar el PDF: {e}")
        return f"Error al generar el PDF: {e}", 500'''

@app.route('/exportar-excel')
def exportar_excel():
    """Generar y exportar un archivo Excel con los registros de asistencia."""
    conexion = conectar_bd()
    if not conexion:
        return "Error al conectar a la base de datos.", 500

    cursor = conexion.cursor(dictionary=True)

    try:
        # Consultar los registros
        cursor.execute("""
            SELECT * from registro
        """)
        registros = cursor.fetchall()

        # Crear un libro de Excel y una hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros de Asistencia"

        # Crear encabezados de columnas
        encabezados = ["ID", "Estudiante", "Curso", "Materia", "Fecha", "Hora", "Estado"]
        ws.append(encabezados)

        # Agregar datos a las filas
        for registro in registros:
            ws.append([
                registro["id_registro"],
                registro["estudiante"],
                registro["curso"],
                registro["materia"],
                registro["fecha"],
                registro["hora"],
                registro["estado"],
                
            ])

        # Guardar el archivo Excel en un buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Enviar el archivo Excel como respuesta
        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=registro_asistencia.xlsx"}
        )
    except Exception as e:
        print(f"Error al generar el archivo Excel: {e}")
        return f"Error al generar el archivo Excel: {e}", 500
    finally:
        cursor.close()
        conexion.close()

@app.route('/exportar-pdf1')
def exportar_pdf1():
    """Generar y exportar un PDF con los registros de asistencia."""
    conexion = conectar_bd()
    if not conexion:
        return "Error al conectar a la base de datos.", 500

    cursor = conexion.cursor(dictionary=True)
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    try:
        # Consultar los registros
        cursor.execute("""
            SELECT id_registro, estudiante, curso, materia, fecha, hora, estado FROM registro
        """)
        registros = cursor.fetchall()

        # Encabezado del PDF
        pdf.cell(200, 10, txt="Reporte de Asistencia1", ln=True, align="C")
        pdf.ln(10)

        # Crear encabezados de columnas
        pdf.set_font("Arial", style="B", size=10)
        encabezados = ["ID", "Estudiante", "Curso ", "Materia", "Fecha", "Hora", "Estado"]
        for encabezado in encabezados:
            pdf.cell(30, 10, txt=encabezado, border=1, align="C")
        pdf.ln()

        # Agregar datos al PDF
        pdf.set_font("Arial", size=9)
        for registro in registros:
            pdf.cell(10, 10, txt=str(registro["id_registro"]), border=1, align="C")
            pdf.cell(40, 10, txt=str(registro["estudiante"]), border=1, align="C")
            pdf.cell(40, 10, txt=str(registro["curso"]), border=1, align="C")
            pdf.cell(30, 10, txt=str(registro["materia"]), border=1, align="C")
            pdf.cell(30, 10, txt=str(registro["fecha"]), border=1, align="C")
            pdf.cell(25, 10, txt=str(registro["hora"]), border=1, align="C")
            pdf.cell(25, 10, txt=str(registro["estado"]), border=1, align="C")
            pdf.ln()

         # Guardar el PDF en un buffer
        output = io.BytesIO()
        pdf.output(dest='S').encode('latin1')  # Guardar el contenido en formato string (latin1)
        output.write(pdf.output(dest='S').encode('latin1'))  # Escribir en el buffer
        output.seek(0)  # Asegurarse de que el puntero esté al principio del buffer

        # Enviar el PDF como respuesta
        return Response(
            output.getvalue(),
            mimetype="application/pdf",
            headers={"Content-Disposition": "inline; filename=registro_asistencia.pdf"}
        )
    except Exception as e:
        print(f"Error al generar el PDF: {e}")
        return f"Error al generar el PDF: {e}", 500
    finally:
        cursor.close()
        conexion.close()

#BD
# Base de datos simulada de estudiantes



# Configuración de la conexión a la base de datos


# Conexión a la base de datos
conn=conectar_bd()
cursor = conn.cursor(dictionary=True)

# Consulta para obtener los datos
query = """
SELECT nombre, foto
FROM estudiante
WHERE foto != '' AND foto IS NOT NULL;
"""
cursor.execute(query)
estudiantes = cursor.fetchall()

# Procesar los datos en la estructura deseada
KNOWN_FACES = []
for estudiante in estudiantes:
    nombre = estudiante['nombre']
    foto = estudiante['foto']
    
    # Construir la ruta de la imagen
    image_path = f"./static/faces/{foto}" if foto != '0' else "./static/faces/default.jpg"
    
    # Agregar a la lista
    KNOWN_FACES.append({
        "name": nombre,
        "image_path": image_path
    })

# Imprimir la lista como JSON para verificar
print(json.dumps(KNOWN_FACES, indent=4))

# Cerrar la conexión
cursor.close()
conn.close()

'''KNOWN_FACES = [
    {
        "name": "Capussiri Alex",
        "image_path": "./static/faces/Alex.jpg"
    },
   {
        "name": "Andrade Jose",
        "image_path": "./static/faces/Jose.jpg"
   },
   {
        "name": "Pedro",
        "image_path": "./static/faces/Pedro.jpg"
   },
   {
        "name": "Zubieta Vania",
        "image_path": "./static/faces/Vania1.jpg"
   }

]'''



#Cargar las imágenes y codificaciones conocidas
known_face_encodings = []
known_face_names = []





#Lista de codificaciones de caras conocidas y nombres
#known_face_encodings = [
#    #Agrega codificaciones aquí, por ejemplo:
#    known_face_encodings(face_recognition.load_image_file("static/faces/Alex.jpg"))[0]
#]
#known_face_names = [
#   "Persona 1"
#]curso

for face in KNOWN_FACES:
    image = face_recognition.load_image_file(face["image_path"])
    encoding = face_recognition.face_encodings(image)[0]
    known_face_encodings.append(encoding)
    known_face_names.append(face["name"])

'''print(known_face_encodings)'''
print(known_face_names)

#@app.route('/control_asistencia', methods=['POST'])
#def control_asistencia():
#    curso_id = request.form.get('curso')
#    materia_id = request.form.get('materia')
#    # Aquí va la lógica para manejar el control de asistencia
#    return f"CONTROL DE ASISTENCIA realizado para Curso {curso_id} y Materia {materia_id}."

'''@app.route('/control_asistencia', methods=['POST'])
def control_asistencia():
    curso = request.form.get('curso')
    materia = request.form.get('materia')
    
    return render_template('control_facial.html', curso=curso, materia=materia)'''

'''@app.route("/control_asistencia/<int:id_curso>/<int:id_materia>")
def control_asistencia(id_curso, id_materia):
    """Página para control de asistencia con transmisión de video y lista de estudiantes."""
    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)

    # Obtener información del curso y materia
    cursor.execute("SELECT nombre_curso FROM curso WHERE id_curso = %s", (id_curso,))
    curso = cursor.fetchone()["nombre_curso"]

    cursor.execute("SELECT nombre_materia FROM materia WHERE id_materia = %s", (id_materia,))
    materia = cursor.fetchone()["nombre_materia"]

    # Obtener lista de estudiantes asociados al curso
    cursor.execute("""
        SELECT id_estudiante, nombre 
        FROM estudiante 
        WHERE curso = %s
    """, (id_curso,))
    estudiantes = cursor.fetchall()

    cursor.close()
    conexion.close()

    return render_template("control_facial.html", curso=curso, materia=materia, estudiantes=estudiantes)'''

@app.route('/control_asistencia', methods=['POST'])
def control_asistencia():
    """Maneja el control de asistencia basado en POST."""
    try:
        # Obtener datos enviados desde el formulario o petición POST
        id_curso = request.form.get('curso')  # Datos enviados desde el frontend
        id_materia = request.form.get('materia')

        # Validar que se recibieron los datos
        if not id_curso or not id_materia:
            return "Error, Faltan datos de curso o materia", 400

        conexion = conectar_bd()
        cursor = conexion.cursor(dictionary=True)

        # Obtener nombre del curso
        cursor.execute("SELECT nombre_curso FROM curso WHERE id_curso = %s", (id_curso,))
        curso_data = cursor.fetchone()
        if not curso_data:
            return "Error: Curso no encontrado", 404
        curso = curso_data["nombre_curso"]

        #curso = cursor.fetchone()["nombre_curso"]

        # Obtener nombre de la materia
        cursor.execute("SELECT nombre_materia FROM materia WHERE id_materia = %s", (id_materia,))
        materia_data = cursor.fetchone()
        if not materia_data:
            return "Error: Materia no encontrada", 404
        materia = materia_data["nombre_materia"]

        # Obtener lista de estudiantes
        cursor.execute("""
            SELECT id_estudiante, nombre
            FROM estudiante 
            WHERE curso = %s
        """, (id_curso,))
        estudiantes = cursor.fetchall()

        cursor.close()
        conexion.close()

        # Renderizar la plantilla con los datos
        return render_template("control_facial.html", curso=curso, materia=materia, estudiantes=estudiantes, id_materia=id_materia, id_curso=id_curso)

    except Exception as e:
        return f"Error al procesar la solicitud: {str(e)}", 500

'''@app.route('/control_asistencia/<int:curso_id>/<int:materia_id>')
def control_asistencia(curso_id, materia_id):
    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)

    # Obtener los datos de la materia seleccionada
    cursor.execute("SELECT id_materia, nombre_materia FROM materia WHERE id_materia = %s", (materia_id,))
    materia = cursor.fetchone()

    # Obtener los estudiantes del curso
    cursor.execute("SELECT id_estudiante, nombre FROM estudiante WHERE curso_id = %s", (curso_id,))
    estudiantes = cursor.fetchall()

    cursor.close()
    conexion.close()

    return render_template('control_asistencia.html', curso=curso_id, materia=materia, estudiantes=estudiantes)'''


@app.route('/video_feed')
def video_feed():
    """RUTA: Transmite el video en vivo."""
    id_materia = request.args.get('id_materia', type=int)  # Obtener id_materia desde la URL
    id_curso = request.args.get('id_curso', type=int)
    if id_materia is None or id_curso is None:
        print("Falta id_materia o id_curso")
        return "Error: Faltan parámetros 'id_materia' o 'id_curso'", 400
    
    return Response(gen_frames(known_face_encodings, known_face_names, id_materia, id_curso), mimetype='multipart/x-mixed-replace; boundary=frame')

def gen_frames1():
    video_capture = cv2.VideoCapture(0)
    video_capture.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    video_capture.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    if not video_capture.isOpened():
        print("No se pudo abrir la cámara.")
        return
    
    try:
        while True:
            success, frame = video_capture.read()
            if not success:
                break

            # Redimensiona para mejorar el rendimiento
            #small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
            if frame is None:
                print("No se está capturando el cuadro correctamente.")

            small_frame = cv2.resize(frame, (0, 0), fx=0.1, fy=0.1)  # Cambia de 0.25 a 0.5

            rgb_small_frame = small_frame[:, :, ::-1]  # Convertir a RGB

            # Detectar rostros en el cuadro reducido
            face_locations = face_recognition.face_locations(rgb_small_frame)
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

            # Escalar las ubicaciones al tamaño original
            scaled_face_locations = [(top * 4, right * 4, bottom * 4, left * 4) 
                                      for (top, right, bottom, left) in face_locations]

            # Dibuja rectángulos y etiquetas en el cuadro original
            for (top, right, bottom, left) in scaled_face_locations:
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                cv2.putText(frame, "Rostro", (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2)

            _, buffer = cv2.imencode('.jpg', frame)
            frame_bytes = buffer.tobytes()

            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
    finally:
        video_capture.release()


def gen_frames4():
    video_capture = cv2.VideoCapture(0)
    if not video_capture.isOpened():
        print("No se pudo abrir la cámara.")
        return

    try:
        while True:
            success, frame = video_capture.read()
            if not success:
                break

            _, buffer = cv2.imencode('.jpg', frame)
            frame_bytes = buffer.tobytes()

            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
    finally:
        video_capture.release()



def gen_frames(known_face_encodings, known_face_names, id_materia,id_curso): 
    """Captura video y realiza reconocimiento facial."""
    print("Iniciando captura de video...")
    estudiantes_diccionario=guardar_estudiantes_en_diccionario()
    materias_diccionario=guardar_materias_en_diccionario()
    video_capture = cv2.VideoCapture(0)

    if not video_capture.isOpened():
        print("No se pudo acceder a la cámara.")
        return

    frame_skip = 5  # Procesa 1 de cada 'frame_skip' cuadros
    frame_count = 0
    
    try:
        while True:
            success, frame = video_capture.read()
            if not success:
                print("Error al leer el cuadro de la cámara.")
                break

            frame_count += 1
            if frame_count % frame_skip != 0:
                continue

            # Reducir el tamaño del cuadro para un procesamiento más rápido
            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
            rgb_small_frame = small_frame[:, :, ::-1]
            rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

            # Detectar rostros y codificar
            face_locations = face_recognition.face_locations(rgb_small_frame, model="hog")
            '''image = face_recognition.load_image_file("static/faces/Alex.jpg")
            face_locations = face_recognition.face_locations(image)'''

            if face_locations:
                print("Se detecto rostro")
                #print(image)
                print(face_locations)
                face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
                
            else:
                face_encodings = []
                print("No se detecto rostro")

            #face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

            # Escalar las ubicaciones al tamaño original
            scaled_face_locations = [(top*4, right*4, bottom*4, left*4) for (top, right, bottom, left) in face_locations]
            asistencia_registrada={}
            for (top, right, bottom, left), face_encoding in zip(scaled_face_locations, face_encodings):
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
                name = "Desconocido"

                if True in matches:
                    first_match_index = matches.index(True)
                    name = known_face_names[first_match_index]
                    # Verificar si el nombre ya está registrado en la sesión
                    ahora = datetime.now().date()
                    if name not in asistencia_registrada or asistencia_registrada[name] != ahora:
                        #obtener datos del estudiante desde el diccionario
                        estudiante=next((est for est in estudiantes_diccionario.values() if est['nombre'] == name), None)
                        #materia=next((mat for mat in materias_diccionario.values()))
                        #materia=materias_diccionario.values()
                        if estudiante:
                            id_estudiante=estudiante['id_estudiante']
                            curso = estudiante['curso']

                            # Verificar que la materia enviada desde el frontend sea válida
                            materia = materias_diccionario.get(id_materia)
                           
                            materia=materia['id_materia']

                            if materia:
                                id_materia_seleccionada = materia
                            else:
                                print(f"Materia con ID {id_materia} no encontrada.")
                                continue
                            # Verificar si ya existe el registro en la base de datos
                            if not verificar_asistencia_existente(id_estudiante, id_curso, id_materia_seleccionada, ahora):
                                guardar_asistencia(id_estudiante, id_curso, id_materia_seleccionada)
                                asistencia_registrada[name] = ahora
                            else:
                                print(f"Asistencia ya registrada para {name} en la fecha {ahora} {id_curso} {materia}")
                            
                            #print(asistencia_registrada)
                        else:
                            print(f"No se encontró información para {name} en el diccionario.")
                    else:
                        print("Rostro no reconocido.")

                # Dibujar un rectángulo alrededor del rostro
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                # Etiquetar con el nombre
                cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2)

            # Codificar la imagen a formato JPEG
            _, buffer = cv2.imencode('.jpg', frame)
            frame_bytes = buffer.tobytes()

            # Enviar el marco como flujo continuo
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
    finally:
        video_capture.release()
        print("Captura de video finalizada.")

def verificar_asistencia_existente(id_estudiante, curso, materia, fecha):
    """Verifica si ya existe un registro de asistencia."""
    conexion = conectar_bd()
    cursor = conexion.cursor()

    consulta = """
        SELECT COUNT(*) 
        FROM registro 
        WHERE estudiante = %s AND curso = %s AND materia = %s AND fecha = %s
    """
    cursor.execute(consulta, (id_estudiante, curso, materia, fecha))
    resultado = cursor.fetchone()

    cursor.close()
    conexion.close()

    return resultado[0] > 0

######################################## CODIGO PARA DOCENTE . funciones ################
# Ojo con esto, solo es para probar
estudiantes = []
asistencia = []
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session or session.get('rol') != 1:
            flash('Acceso denegado: solo para administradores', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def docente_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session or session.get('rol') != 2:
            flash('Acceso denegado: solo para docentes', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def padre_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session or session.get('rol') != 3:
            flash('Acceso denegado: solo para padres', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/dashboard_admin')
@admin_required  # Asegúrate de tener decorador para validar roles
def dashboard_admin():
    return render_template('dashboard_admin.html', user=session.get('usuario_id'))

@app.route('/dashboard_docente')
@docente_required  # Decorador para validar docente
def dashboard_docente():
    return render_template('dashboard_docente.html', user=session.get('usuario_id'))

#@app.route('/dashboard_padre')

#def dashboard_padre():
   # return render_template('dashboard_padre.html', user=session.get('usuario_id'))
                           



@app.route('/generar_excel')
def generar_excel():
    df = pd.DataFrame(asistencia)
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Asistencia')
    writer.save()
    output.seek(0)
    return send_file(output, attachment_filename="asistencia.xlsx", as_attachment=True)

@app.route('/generar_pdf')
def generar_pdf():
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Registro de Asistencia", ln=True, align='C')
    
    # Encabezados
    pdf.cell(30, 10, "CI", 1)
    pdf.cell(50, 10, "Apellidos", 1)
    pdf.cell(50, 10, "Nombre", 1)
    pdf.cell(30, 10, "Curso", 1)
    pdf.cell(30, 10, "Materia", 1)
    pdf.ln()
    
    # Datos
    for registro in asistencia:
        pdf.cell(30, 10, registro["CI"], 1)
        pdf.cell(50, 10, registro["Apellidos"], 1)
        pdf.cell(50, 10, registro["Nombre"], 1)
        pdf.cell(30, 10, registro["Curso"], 1)
        pdf.cell(30, 10, registro["Materia"], 1)
        pdf.ln()
    
    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return send_file(output, attachment_filename="asistencia.pdf", as_attachment=True)

@app.route('/capture_attendance', methods=['GET'])
def capture_attendance():
    recognizer = FaceRecognition(faces_path="D:/Documentos/4_Maestria/Proyecto2/proyecto/static/faces")
    recognizer.recognize_faces()
    return redirect(url_for('dashboard'))


@app.route("/reportes")
def reportes():
    """Ruta para mostrar la página de reportes."""
    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)

    # Obtener lista de materias y cursos
    cursor.execute("SELECT id_materia, nombre_materia FROM materia")
    materias = cursor.fetchall()

    cursor.execute("SELECT id_curso, nombre_curso FROM curso")
    cursos = cursor.fetchall()

    cursor.close()
    conexion.close()

    return render_template("reportes.html", materias=materias, cursos=cursos)


from fpdf import FPDF
import pandas as pd

@app.route("/generar_reporte1", methods=["POST"])
def generar_reporte1():
    """Genera el reporte de asistencia basado en filtros."""
    try:
        # Obtener datos del formulario
        materia_id = request.form.get("materia")
        curso_id = request.form.get("curso")
        fecha_inicio = request.form.get("fecha_inicio")
        fecha_fin = request.form.get("fecha_fin")

        # Validar fechas
        fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        fecha_fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()

        conexion = conectar_bd()
        cursor = conexion.cursor(dictionary=True)

        # Consulta para obtener los registros filtrados
        consulta = """
        SELECT r.id_registro, e.nombre, c.nombre_curso, m.nombre_materia, r.fecha, r.hora, r.estado
        FROM registro r
        INNER JOIN estudiante e ON r.estudiante = e.id_estudiante
        INNER JOIN curso c ON e.curso = c.id_curso
        INNER JOIN materia m ON r.materia = m.id_materia
        WHERE r.materia = %s AND c.id_curso = %s AND r.fecha BETWEEN %s AND %s
        ORDER BY r.fecha, r.hora
        """

        cursor.execute(consulta, (materia_id, curso_id, fecha_inicio, fecha_fin))
        registros = cursor.fetchall()

        cursor.close()
        conexion.close()

        return render_template("reporte_resultados.html", registros=registros, 
                               materia_id=materia_id, curso_id=curso_id, 
                               fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


@app.route("/generar_reporte2", methods=["POST"])
def generar_reporte2():
    """Genera el reporte de asistencia mostrando todos los estudiantes del curso y su estado."""
    try:
        # Obtener datos del formulario
        materia_id = request.form.get("materia")
        curso_id = request.form.get("curso")
        fecha_inicio = request.form.get("fecha_inicio")
        fecha_fin = request.form.get("fecha_fin")

        # Validar fechas
        fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        fecha_fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()

        conexion = conectar_bd()
        cursor = conexion.cursor(dictionary=True)

        # Consulta para obtener los registros filtrados
        consulta = """
        SELECT 
            e.id_estudiante,
            e.nombre,
            c.nombre_curso,
            m.nombre_materia,
            r.fecha,
            r.hora,
            IF(r.estado IS NOT NULL, 'Asistido', 'Falta') AS estado
        FROM estudiante e
        INNER JOIN curso c ON e.curso = c.id_curso
        INNER JOIN materia m ON m.id_materia = %s
        LEFT JOIN registro r ON e.id_estudiante = r.estudiante 
            AND r.materia = m.id_materia 
            AND r.fecha BETWEEN %s AND %s
        WHERE e.curso = %s
        ORDER BY e.nombre, r.fecha, r.hora
        """

        cursor.execute(consulta, (materia_id, fecha_inicio, fecha_fin, curso_id))
        registros = cursor.fetchall()

        cursor.close()
        conexion.close()

        return render_template("reporte_resultados.html", registros=registros, 
                               materia_id=materia_id, curso_id=curso_id, 
                               fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


import tempfile

from flask import send_file, jsonify
from fpdf import FPDF
import tempfile

@app.route("/descargar_pdf1", methods=["GET"])
def descargar_pdf1():
    try:
        # Obtener los filtros enviados desde la tabla
        materia_id = request.args.get("materia_id")
        curso_id = request.args.get("curso_id")
        fecha_inicio = request.args.get("fecha_inicio")
        fecha_fin = request.args.get("fecha_fin")

        conexion = conectar_bd()
        cursor = conexion.cursor(dictionary=True)

        # 🔹 Consulta para incluir a TODOS los estudiantes del curso, asignando "Falta" a los que no tienen registro
        consulta = """
        SELECT 
            e.nombre, 
            c.nombre_curso, 
            m.nombre_materia, 
            COALESCE(r.fecha, '') AS fecha, 
            COALESCE(r.hora, '') AS hora, 
            COALESCE(r.estado, 'Falta') AS estado
        FROM estudiante e
        INNER JOIN curso c ON e.curso = c.id_curso
        INNER JOIN materia m ON m.id_materia = %s
        LEFT JOIN registro r ON e.id_estudiante = r.estudiante 
            AND r.materia = m.id_materia 
            AND r.fecha BETWEEN %s AND %s
        WHERE e.curso = %s
        ORDER BY e.nombre, r.fecha, r.hora
        """

        cursor.execute(consulta, (materia_id, fecha_inicio, fecha_fin, curso_id))
        registros = cursor.fetchall()

        cursor.close()
        conexion.close()

        if not registros:
            return jsonify({"status": "error", "message": "No hay registros para generar el PDF."})

        # 🔹 Generar el PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        pdf.cell(200, 10, txt="Reporte de Asistencia", ln=True, align='C')
        pdf.ln(10)

        # Encabezados de la tabla
        pdf.cell(50, 10, "Estudiante", 1)
        pdf.cell(20, 10, "Curso", 1)
        pdf.cell(30, 10, "Materia", 1)
        pdf.cell(30, 10, "Fecha", 1)
        pdf.cell(30, 10, "Hora", 1)
        pdf.cell(30, 10, "Estado", 1)
        pdf.ln()

        # 🔹 Agregar registros a la tabla
        for registro in registros:
            pdf.cell(50, 10, str(registro["nombre"]), 1)
            pdf.cell(20, 10, str(registro["nombre_curso"]), 1)
            pdf.cell(30, 10, str(registro["nombre_materia"]), 1)
            pdf.cell(30, 10, str(registro["fecha"]), 1)  
            pdf.cell(30, 10, str(registro["hora"]), 1)   
            pdf.cell(30, 10, str(registro["estado"]), 1) 
            pdf.ln()

        # Guardar el PDF temporalmente y enviarlo al usuario
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf:
            pdf.output(tmp_pdf.name)
            return send_file(tmp_pdf.name, as_attachment=True, download_name="reporte_asistencia.pdf")

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


import pandas as pd
from flask import send_file, request, jsonify
import tempfile

@app.route("/descargar_excel1", methods=["GET"])
def descargar_excel1():
    try:
        # Obtener los filtros desde la URL
        materia_id = request.args.get("materia_id")
        curso_id = request.args.get("curso_id")
        fecha_inicio = request.args.get("fecha_inicio")
        fecha_fin = request.args.get("fecha_fin")

        if not materia_id or not curso_id or not fecha_inicio or not fecha_fin:
            return jsonify({"status": "error", "message": "Faltan parámetros para generar el reporte."})

        conexion = conectar_bd()
        cursor = conexion.cursor(dictionary=True)

        # 🔹 Consulta mejorada: Asegura que todos los estudiantes aparezcan, asignando "Falta" cuando no hay asistencia
        consulta = """
        SELECT 
            e.nombre AS Estudiante, 
            c.nombre_curso AS Curso, 
            m.nombre_materia AS Materia, 
            COALESCE(r.fecha, '') AS Fecha, 
            COALESCE(r.hora, '') AS Hora, 
            COALESCE(r.estado, 'Falta') AS Estado
        FROM estudiante e
        INNER JOIN curso c ON e.curso = c.id_curso
        INNER JOIN materia m ON m.id_materia = %s
        LEFT JOIN registro r ON e.id_estudiante = r.estudiante 
            AND r.materia = m.id_materia 
            AND (r.fecha BETWEEN %s AND %s OR r.fecha IS NULL)
        WHERE e.curso = %s
        ORDER BY e.nombre, r.fecha, r.hora
        """

        cursor.execute(consulta, (materia_id, fecha_inicio, fecha_fin, curso_id))
        registros = cursor.fetchall()

        cursor.close()
        conexion.close()

        # 🔹 Depuración: Mostrar cuántos registros se obtuvieron
        if not registros:
            return jsonify({"status": "error", "message": "No hay registros para generar el Excel."})

        # Convertir registros en un DataFrame de pandas
        df = pd.DataFrame(registros)

        # 🔹 Guardar el archivo Excel temporalmente y enviarlo
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
            df.to_excel(tmp_excel.name, index=False, sheet_name="Asistencia")
            return send_file(tmp_excel.name, as_attachment=True, download_name="reporte_asistencia.xlsx")

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})



    
@app.route('/dashboard_padre')
@padre_required  # Asegúrate de tener decorador para validar roles
def dashboard_padre():
    print("⚡⚡⚡ EJECUTANDO /generar_reporte ⚡⚡⚡")
    print("Sesión actual:", session)
    '''if 'usuario_id' not in session or session.get('rol') != 'Padre':
        print("❌ No hay usuario autenticado o no es un padre.")
        flash('Acceso no autorizado', 'danger')
        return redirect(url_for('login'))'''

    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)
    
    padre_id = session['usuario_id']

    # Obtener hijos del padre autenticado
    padre_id = session['usuario_id']
    print("✅ Padre autenticado, ID:", padre_id)
    cursor.execute("""
        SELECT e.id_estudiante, e.nombre 
        FROM estudiante e
        WHERE e.id_usuario = %s
    """, (padre_id,))
    hijos = cursor.fetchall()
    #hijos = [{"id_estudiante": fila[0], "nombre": fila[1]} for fila in cursor.fetchall()]
    print("👦 Hijos obtenidos:", hijos)
    materias = []
    reporte = []

    hijo_seleccionado = request.args.get('hijo')
    materia_seleccionada = request.args.get('materia')
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')

    if hijo_seleccionado:
        print("Hijo seleccionado:", hijo_seleccionado)
        cursor.execute("""
            SELECT m.id_materia, m.nombre_materia 
                       FROM materia m 
                       JOIN curso_materia cm ON m.id_materia = cm.materia 
                       JOIN curso c ON cm.curso = c.id_curso 
                       JOIN estudiante e ON c.id_curso=e.curso 
                       WHERE e.id_estudiante = %s
        """, (hijo_seleccionado,))
        materias = cursor.fetchall()
        #materias = [{"id_materia": fila[0], "nombre_materia": fila[1]} for fila in cursor.fetchall()]
        print("📌 Materias obtenidas:", materias)
    if hijo_seleccionado and fecha_inicio and fecha_fin:
            consulta = """
                SELECT r.fecha, r.hora, e.nombre, m.nombre_materia AS materia, r.estado 
                FROM registro r 
                JOIN materia m ON r.materia = m.id_materia
                JOIN estudiante e ON r.estudiante=e.id_estudiante 
                WHERE r.estudiante = %s AND r.fecha BETWEEN %s AND %s
            """
            parametros = [hijo_seleccionado, fecha_inicio, fecha_fin]

            if materia_seleccionada:
                consulta += " AND r.materia = %s"
                parametros.append(materia_seleccionada)

            cursor.execute(consulta, tuple(parametros))
            reporte = cursor.fetchall()

    conexion.close()

    return render_template('dashboard_padre.html', user=session.get('usuario_id'), hijos=hijos, materias=materias, reporte=reporte)

@app.route('/descargar_reporte_pdf')
def descargar_reporte_pdf():
    hijo_id = request.args.get('hijo')
    materia_id = request.args.get('materia')
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')

    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)

    # Consulta para obtener los datos del reporte
    consulta = """
        SELECT r.fecha, r.hora, e.nombre, m.nombre_materia AS materia, r.estado 
                FROM registro r 
                JOIN materia m ON r.materia = m.id_materia
                JOIN estudiante e ON r.estudiante=e.id_estudiante 
                WHERE r.estudiante = %s AND r.fecha BETWEEN %s AND %s
    """
    parametros = [hijo_id, fecha_inicio, fecha_fin]

    if materia_id:
        consulta += " AND r.materia = %s"
        parametros.append(materia_id)

    cursor.execute(consulta, tuple(parametros))
    reporte = cursor.fetchall()
    conexion.close()

    # Crear PDF en memoria
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    pdf.setTitle("Reporte de Asistencia")

    pdf.drawString(200, 750, "Reporte de Asistencia")
    pdf.drawString(100, 730, f"Hijo ID: {hijo_id}")

    y = 700
    pdf.drawString(100, y, "Nombre   |  Materia   |  Fecha       | Materia      | Estado")
    pdf.line(100, y - 5, 500, y - 5)

    for fila in reporte:
        y -= 20
        pdf.drawString(100, y, f"{fila['nombre']} | {fila['materia']} | {fila['fecha']}  | {fila['hora']}  | {fila['estado']}")

    pdf.showPage()
    pdf.save()
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name="reporte_asistencia.pdf", mimetype='application/pdf')


@app.route('/descargar_reporte_excel')
def descargar_reporte_excel():
    hijo_id = request.args.get('hijo')
    materia_id = request.args.get('materia')
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')

    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)

    # Consulta para obtener los datos del reporte
    consulta = """
        SELECT r.fecha, r.hora, e.nombre, m.nombre_materia AS materia, r.estado 
                FROM registro r 
                JOIN materia m ON r.materia = m.id_materia
                JOIN estudiante e ON r.estudiante=e.id_estudiante 
                WHERE r.estudiante = %s AND r.fecha BETWEEN %s AND %s
    """
    parametros = [hijo_id, fecha_inicio, fecha_fin]

    if materia_id:
        consulta += " AND r.materia = %s"
        parametros.append(materia_id)

    cursor.execute(consulta, tuple(parametros))
    reporte = cursor.fetchall()
    conexion.close()

    # Convertir los datos a DataFrame
    df = pd.DataFrame(reporte)
    
    # Guardar en un archivo en memoria
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Asistencia")

    output.seek(0)

    return send_file(output, as_attachment=True, download_name="reporte_asistencia.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/reporte_estudiante', methods=['GET', 'POST'])
def reporte_estudiante():
    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)

    if request.method == 'POST':
        # Obtener datos del formulario
        nombre_estudiante = request.form['nombre_estudiante']
        fecha_inicio = request.form['fecha_inicio']
        fecha_fin = request.form['fecha_fin']

        # Consulta para buscar registros por estudiante y rango de fechas
        query = """
        SELECT registro.id_registro, estudiante.nombre, curso.nombre_curso, materia.nombre_materia, 
               registro.fecha, registro.hora, registro.estado
        FROM registro
        INNER JOIN estudiante ON registro.estudiante = estudiante.id_estudiante
        INNER JOIN curso ON registro.curso = curso.id_curso
        INNER JOIN materia ON registro.materia = materia.id_materia
        WHERE (estudiante.nombre LIKE %s)
          AND registro.fecha BETWEEN %s AND %s
        """
        cursor.execute(query, (f"%{nombre_estudiante}%", f"%{nombre_estudiante}%", fecha_inicio, fecha_fin))
        registros = cursor.fetchall()
        return render_template('reporte_resultados1.html', registros=registros)

    # Renderizar la página inicial del formulario
    return render_template('reporte_estudiante.html')


@app.route('/descargar_pdf', methods=['POST'])
def descargar_pdf():
    estudiante = request.form.get('estudiante')
    fecha_inicio = request.form.get('fecha_inicio')
    fecha_fin = request.form.get('fecha_fin')

    # Construir la consulta
    query = """
        SELECT r.id_registro, e.nombres, e.apellidos, c.nombre_curso AS curso, m.nombre_materia AS materia, r.fecha, r.estado
        FROM registro r
        INNER JOIN estudiante e ON r.estudiante = e.id_estudiante
        INNER JOIN curso c ON e.curso = c.id_curso
        INNER JOIN materia m ON r.materia = m.id_materia
        WHERE 1=1
    """
    filters = []

    if estudiante:
        query += " AND e.nombre LIKE %s"
        filters.append(f"%{estudiante}%")
    if fecha_inicio:
        query += " AND r.fecha >= %s"
        filters.append(fecha_inicio)
    if fecha_fin:
        query += " AND r.fecha <= %s"
        filters.append(fecha_fin)

    # Ejecutar la consulta
    try:
        conexion = conectar_bd()
        cursor = conexion.cursor(dictionary=True)
        cursor.execute(query, tuple(filters))
        registros = cursor.fetchall()  # Asegurarse de que se está llamando fetchall() en una consulta SELECT
        cursor.close()
    except Exception as e:
        print(f"Error al ejecutar la consulta: {e}")
        return "Error al generar el reporte", 500

    # Verificar si hay resultados
    if not registros:
        return "No se encontraron registros para los criterios dados.", 404

    # Generar el PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Reporte de Asistencia", ln=True, align="C")

    # Agregar encabezado
    pdf.cell(20, 10, "ID", border=1)
    pdf.cell(50, 10, "Nombres", border=1)
    pdf.cell(50, 10, "Apellidos", border=1)
    pdf.cell(30, 10, "Curso", border=1)
    pdf.cell(40, 10, "Materia", border=1)
    pdf.cell(30, 10, "Fecha", border=1)
    pdf.cell(20, 10, "Estado", border=1)
    pdf.ln()

    # Agregar registros al PDF
    for registro in registros:
        pdf.cell(20, 10, str(registro['id_registro']), border=1)
        pdf.cell(50, 10, registro['nombres'], border=1)
        pdf.cell(50, 10, registro['apellidos'], border=1)
        pdf.cell(30, 10, registro['curso'], border=1)
        pdf.cell(40, 10, registro['materia'], border=1)
        pdf.cell(30, 10, registro['fecha'].strftime('%Y-%m-%d'), border=1)
        pdf.cell(20, 10, registro['estado'], border=1)
        pdf.ln()

    # Preparar la respuesta
    pdf_output = pdf.output(dest='S').encode('latin1')
    response = make_response(pdf_output)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=reporte_asistencia.pdf'
    return response

############### AGREGAR ESTUDIANTE ########################
# Configuración del sistema
UPLOAD_FOLDER = 'static/faces'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Función para validar extensiones permitidas
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Ruta para mostrar el formulario
@app.route('/agregar_estudiante', methods=['GET', 'POST'])
def agregar_estudiante():
    cursos = obtener_cursos()  # Función para obtener cursos desde la base de datos
    #materias = obtener_materias1()  # Función para obtener materias desde la base de datos
    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)
    # Obtener lista de padres
    cursor.execute("SELECT id_usuario, nombre FROM usuario WHERE id_rol = 3")  # 3 es el rol de padres
    padres = cursor.fetchall()

    if request.method == 'POST':
        # Obtener los datos del formulario
        nombre = request.form['nombre']
        curso = request.form['curso']
        ci = request.form['ci']
        #materia = request.form['materia']
        id_usuario = request.form['id_padre']
        foto = request.files['foto']

        if foto and allowed_file(foto.filename):
            # Guardar la fotografía con un nombre seguro
            filename = secure_filename(foto.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            foto.save(filepath)

            conexion = None
            cursor = None

            # Guardar la foto
            #with open(filepath, 'wb') as f:
             #   f.write(foto.read())

            # Codificar el rostro para el reconocimiento facial
            try:
                image = face_recognition.load_image_file(filepath)
                face_encodings = face_recognition.face_encodings(image)
                if len(face_encodings) > 0:
                    # Conectar a la base de datos y guardar la información
                    conexion=conectar_bd()
                    cursor = conexion.cursor()
                    consulta = """
                        INSERT INTO estudiante (CI, nombre, curso, foto, id_usuario)
                        VALUES (%s, %s, %s, %s, %s)
                    """
                    cursor.execute(consulta, (ci, nombre, curso, filename, id_usuario))
                    conexion.commit()
                    
                    flash("Estudiante agregado exitosamente.", "success")
                    print("Estudiante agregado exitosamente.", "success")
                else:
                    flash("No se detectó un rostro en la fotografía. Intenta con otra imagen.", "danger")
                    os.remove(filepath)
                
                print(f"Datos a insertar: CI={ci}, Nombre={nombre}, Curso={curso}, Email={id_usuario}, Foto={filename}")

            except Exception as e:
                flash(f"Error al procesar la fotografía: {e}", "danger")
                os.remove(filepath)
                print(f"Datos a insertar: CI={ci}, Nombre={nombre}, Curso={curso}, Email={id_usuario}, Foto={filename}")
            finally:
                # Asegúrate de cerrar el cursor y la conexión solo si están definidos
                if cursor:
                    cursor.close()
                if conexion:
                    conexion.close()
        else:
            flash("Archivo no permitido. Sube una imagen en formato PNG, JPG o JPEG.", "danger")

        return redirect(url_for('dashboard'))

    return render_template('agregar_estudiante.html', cursos=cursos, padres=padres)

@app.route('/buscar_padres')
def buscar_padres():
    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)
    termino = request.args.get('q', '')
    
    query = "SELECT id_usuario, nombre FROM usuario WHERE id_rol = 3 AND nombre LIKE %s LIMIT 10"
    cursor.execute(query, (f"%{termino}%",))
    resultados = cursor.fetchall()
    
    conexion.close()
    return jsonify(resultados)

############# AGREGAR DOCENTES ###############
# Configuración de la base de datos (ajústala según tu configuración)
db= conectar_bd()

cursor = db.cursor(dictionary=True)

# Middleware para verificar si el usuario es administrador
'''def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario' not in session or session.get('rol') != 'administrador':
            return jsonify({"error": "Acceso denegado"}), 403
        return f(*args, **kwargs)
    return decorated_function'''

# Ruta para agregar un docente
@app.route('/agregar_usuario', methods=['POST','GET'])
#@admin_required
def agregar_usuario():
    if request.method == 'POST':
        conexion=conectar_bd()
        cursor = conexion.cursor()
        #data = request.json
        nombre = request.form.get('nombre')
        correo = request.form.get('correo')
        password= request.form.get('password')
        materia = request.form.get('materia')
        rol = request.form.get('rol')
        ########### PARA USUARIO ADMIN ##########
        '''cursor.execute("SELECT * FROM usuario WHERE correo = %s", (correo,))
        usuario = cursor.fetchone()
        if not usuario:  # Si el usuario no existe, lo creamos (solo para pruebas)
            nombre = "admin"
            correo1 = "admin@admin.com"
            contrasena1 = "admin123"
            hash_contrasena = generate_password_hash(contrasena1)

            cursor.execute("INSERT INTO usuario (nombre, correo, contrasena, id_rol) VALUES (%s, %s, %s, %s)", 
                           (nombre, correo1, hash_contrasena, 1))  
            conexion.commit()
            flash('Usuario admin creado. Intenta iniciar sesión.', 'info')
            return redirect(url_for('login'))
        conexion.close()'''
        ############# USUARIO ADMIN #############




        

            # Hash de la contraseña
        hashed_password = generate_password_hash(password)
            
        if rol == "2":  # Si el rol es "Docente"
            if not materia:  # Validar que se seleccionó una materia
                flash("Debe seleccionar una materia para el docente", "danger")
                return redirect(url_for('agregar_usuario'))
            
            cursor.execute(
                "INSERT INTO usuario (nombre, correo, contrasena, id_materia, id_rol) VALUES (%s, %s, %s, %s, %s)",
                (nombre, correo, hashed_password, materia, rol)
            )
        else:
            # Si NO es docente, se inserta sin materia
            cursor.execute(
                "INSERT INTO usuario (nombre, correo, contrasena, id_rol) VALUES (%s, %s, %s, %s)",
                (nombre, correo, hashed_password, rol)
            )
        conexion.commit() 
                
        flash("Usuario agregado con éxito.", "success")
        return redirect(url_for('dashboard'))  
                    

     # Obtener las materias y roles
    materias, roles = [], []
    try:
        conexion = conectar_bd()
        cursor = conexion.cursor()
        cursor.execute("SELECT id_materia, nombre_materia FROM materia")
        materias = [{"id_materia": fila[0], "nombre_materia": fila[1]} for fila in cursor.fetchall()]

        cursor.execute("SELECT id_rol, nombre FROM rol")
        roles = [{"id_rol": fila[0], "nombre": fila[1]} for fila in cursor.fetchall()]

    except mysql.connector.Error as err:
        flash(f"Error al obtener datos: {str(err)}", "danger")
    finally:
        if conexion:
            conexion.close()

    return render_template('agregar_usuario.html', materias=materias, roles=roles)
               
############# AGREGAR MATERIA ##############
@app.route('/agregar_materia', methods=['POST','GET'])
#@admin_required
def agregar_materia():
    if request.method == 'POST':
        conexion=conectar_bd()
        cursor = conexion.cursor()
        #data = request.json
        nombre = request.form.get('nombre')
               
        if not nombre:
            flash("El nombre de la materia es obligatorio.", "danger")
            return redirect(url_for('agregar_materia'))
        try:
            cursor.execute("INSERT INTO materia (nombre_materia) VALUES (%s)", (nombre,))
            conexion.commit()
            flash("Materia agregada con éxito.", "success")
        except Exception as e:
            flash(f"Error al agregar la materia: {str(e)}", "danger")
        finally:
            cursor.close()
            conexion.close()

        return redirect(url_for('agregar_materia'))  # Redirige para evitar reinserción en recarga

    return render_template('agregar_materia.html')
# ########## FIN DE AGREGAR MATERIA ########                      

############# AGREGAR CURSO ##############
'''@app.route('/agregar_curso', methods=['POST','GET'])
#@admin_required
def agregar_curso():
    if request.method == 'POST':
        conexion=conectar_bd()
        cursor = conexion.cursor()
        #data = request.json
        nombre = request.form.get('nombre')
               
        if not nombre:
            flash("El nombre del curso es obligatorio.", "danger")
            return redirect(url_for('agregar_curso'))
        try:
            cursor.execute("INSERT INTO curso (nombre_curso) VALUES (%s)", (nombre,))
            conexion.commit()
            flash("Curso agregado con éxito.", "success")
        except Exception as e:
            flash(f"Error al agregar el curso: {str(e)}", "danger")
        finally:
            cursor.close()
            conexion.close()

        return redirect(url_for('agregar_curso'))  # Redirige para evitar reinserción en recarga

    return render_template('agregar_curso.html')
# ########## FIN DE AGREGAR curso ########'''

@app.route('/agregar_curso', methods=['POST', 'GET'])
def agregar_curso():
    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)

    # Obtener todas las materias disponibles
    cursor.execute("SELECT id_materia, nombre_materia FROM materia")

    materias = cursor.fetchall()
    #materias = [{"id_materia": fila[0], "nombre_materia": fila[1]} for fila in cursor.fetchall()]
    
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        materias_seleccionadas = request.form.getlist('materias')  # Obtiene una lista de materias seleccionadas

        if not nombre:
            flash("El nombre del curso es obligatorio.", "danger")
            return redirect(url_for('agregar_curso'))

        try:
            # Insertar el nuevo curso
            cursor.execute("INSERT INTO curso (nombre_curso) VALUES (%s)", (nombre,))
            conexion.commit()

            # Obtener el ID del curso recién insertado
            id_curso = cursor.lastrowid  

            # Insertar las materias seleccionadas en la tabla intermedia curso_materia
            for id_materia in materias_seleccionadas:
                cursor.execute("INSERT INTO curso_materia (curso, materia) VALUES (%s, %s)", (id_curso, id_materia))
            conexion.commit()

            flash("Curso agregado con éxito.", "success")
        except Exception as e:
            flash(f"Error al agregar el curso: {str(e)}", "danger")
        finally:
            cursor.close()
            conexion.close()

        return redirect(url_for('agregar_curso'))

    return render_template('agregar_curso.html', materias=materias)
                   


###################  ##################################
# Cargar rostros conocidos desde la base de datos
def cargar_rostros_conocidos():
    conexion=conectar_bd()
    cursor = conexion.cursor(dictionary=True)

    consulta = "SELECT id_estudiante, nombre, foto FROM estudiante"
    cursor.execute(consulta)
    estudiantes = cursor.fetchall()

    known_face_encodings = []
    known_face_names = []

    for estudiante in estudiantes:
        filepath = os.path.join(UPLOAD_FOLDER, estudiante['foto'])
        image = face_recognition.load_image_file(filepath)
        face_encodings = face_recognition.face_encodings(image)
        if len(face_encodings) > 0:
            known_face_encodings.append(face_encodings[0])
            known_face_names.append(estudiante['nombre'])

    cursor.close()
    conexion.close()

    return known_face_encodings, known_face_names

@app.route('/obtener_materias/<int:id_curso>')
def obtener_materias(id_curso):
    """Devuelve las materias relacionadas con un curso en formato JSON."""
    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)

    consulta = """
        SELECT m.id_materia, m.nombre_materia 
        FROM curso_materia cm
        INNER JOIN materia m ON cm.materia = m.id_materia
        WHERE cm.curso = %s
    """
    cursor.execute(consulta, (id_curso,))
    materias = cursor.fetchall()

    cursor.close()
    conexion.close()

    return jsonify(materias)

'''@app.route("/obtener_materias/<int:curso_id>", methods=["GET"])
def obtener_materias(curso_id):
    """Obtiene las materias asociadas a un curso específico."""
    try:
        conexion = conectar_bd()
        cursor = conexion.cursor(dictionary=True)

        consulta = "SELECT id_materia, nombre_materia FROM materia WHERE curso_id = %s"
        cursor.execute(consulta, (curso_id,))
        materias = cursor.fetchall()

        cursor.close()
        conexion.close()

        return jsonify(materias)
    except Exception as e:
        return jsonify({"error": str(e)})'''


############### LISTAR DOCENTES ###############
@app.route('/listar_docentes', methods=['GET'])
def listar_docentes():
    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)

    consulta = """
    SELECT DISTINCT u.id_usuario, u.nombre, u.correo, u.contrasena, m.id_materia, m.nombre_materia
    FROM usuario u
    JOIN materia m ON u.id_materia = m.id_materia;
    """
    
    cursor.execute(consulta)
    docentes = cursor.fetchall()

    cursor.execute("SELECT id_materia, nombre_materia FROM materia")
    materias = cursor.fetchall()

    #cursor.close()  
    conexion.close()
    #return jsonify(docentes)
    return render_template('listar_docentes.html', docentes=docentes, materias=materias)

@app.route('/editar_docente/<int:id>', methods=['POST'])
def editar_docente(id):
    data = request.json
    conexion = conectar_bd()
    cursor = conexion.cursor()
    if not data:
        return jsonify({'error': 'No se recibieron datos'}), 400
    
    nombre = data.get('nombre')
    correo = data.get('correo')
    id_materia = data.get('materia')

    if not all([nombre, correo, id_materia]):
        return jsonify({'error': 'Todos los campos son obligatorios'}), 400

    try:
        consulta = """UPDATE usuario 
                      SET nombre=%s, correo=%s, id_materia=%s 
                      WHERE id_usuario=%s"""
        cursor.execute(consulta, (nombre, correo, id_materia, id))
        conexion.commit()
        return jsonify({'mensaje': 'Docente actualizado correctamente'})
    except Exception as e:
        return jsonify({'error': f'Error al actualizar docente: {str(e)}'}), 500
    finally:
        conexion.close()

@app.route('/eliminar_docente/<int:id>', methods=['POST'])
def eliminar_docente(id):
    conexion = conectar_bd()
    cursor = conexion.cursor()
    try:
        cursor.execute("DELETE FROM usuario WHERE id_usuario=%s", (id,))
        conexion.commit()
        return jsonify({'message': 'Docente eliminado correctamente'})
    except mysql.connector.Error as err:
        return jsonify({'error': str(err)}), 500
    finally:
        cursor.close()
        conexion.close()

# Ruta para obtener datos de un docente por ID
@app.route('/obtener_docente/<int:id>', methods=['GET'])
def obtener_docente(id):
    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)
    
    cursor.execute("SELECT * FROM usuario WHERE id_usuario = %s", (id,))
    docente = cursor.fetchone()

    conexion.close()
    return jsonify(docente)
#@app.route('/docentes')
#def docentes():
 #   return render_template('listar_docentes.html')
############# FIN DE LISTAR DOCENTES ###########

##### EDITAR ESTADO ##########
@app.route("/actualizar_estado_asistencia", methods=["POST"])
def actualizar_estado_asistencia():
    try:
        data = request.get_json()
        id_registro = data.get("id_registro")
        nuevo_estado = data.get("estado")

        conexion = conectar_bd()
        cursor = conexion.cursor()
        
        consulta = """
        UPDATE registro
        SET estado = %s
        WHERE id_registro = %s
        """
        cursor.execute(consulta, (nuevo_estado, id_registro))
        conexion.commit()
        
        cursor.close()
        conexion.close()
        
        return jsonify({"status": "success", "message": "Estado actualizado correctamente"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

##### FIN EDITAR ESTADO #######
########## MIS HIJOS ###############
@app.route('/mis_hijos')
def mis_hijos():
    if 'usuario_id' not in session or session['rol'] != 'Padre de familia':
        return redirect(url_for('login'))

    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)
    
    cursor.execute("""
        SELECT e.id_estudiante, e.CI, e.nombre, e.curso, e.email_ppff
        FROM estudiante e
        JOIN padre_estudiante pe ON e.id_estudiante = pe.id_estudiante
        JOIN padre_familia pf ON pe.id_padre = pf.id_padre
        JOIN usuario u ON pf.id_usuario = u.id_usuario
        WHERE u.id_usuario = %s
    """, (session['usuario_id'],))

    hijos = cursor.fetchall()
    
    return render_template('mis_hijos.html', hijos=hijos)

########## FIN DE MIS HIJOS ########
############## REPORTE DE ASISTENCIA DE LOS HIJOS ###############
@app.route('/reporte_asistencia/<int:id_estudiante>')
def reporte_asistencia(id_estudiante):
    if 'usuario_id' not in session or session['rol'] != 'Padre de familia':
        return redirect(url_for('login'))

    conexion = conectar_bd()
    cursor = conexion.cursor(dictionary=True)

    # Verificar que el estudiante realmente es hijo del padre de familia
    cursor.execute("""
        SELECT * FROM padre_estudiante pe
        JOIN padre p ON pe.id_padre = p.id_padre
        WHERE p.id_usuario = %s AND pe.id_estudiante = %s
    """, (session['usuario_id'], id_estudiante))

    relacion = cursor.fetchone()
    if not relacion:
        flash("No tienes permiso para ver este reporte.", "danger")
        return redirect(url_for('mis_hijos'))

    # Obtener la asistencia del estudiante
    cursor.execute("""
        SELECT fecha, estado FROM registro WHERE estudiante = %s
    """, (id_estudiante,))
    
    asistencias = cursor.fetchall()
    
    return render_template('reporte_asistencia.html', asistencias=asistencias)

###############FIN DE REPORTE DE ASISTENCIA DE LOS HIJOS#########

if __name__ == '__main__':
    print("🚀 Servidor Flask y Scheduler de notificaciones iniciados...")
    app.run(debug=True)

